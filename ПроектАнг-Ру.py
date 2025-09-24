## -- Импорт библиотек и таблицы

import openpyxl
import random
from tkinter import *
import tkinter as tk
from tkinter import ttk, END
from tkinter.messagebox import showinfo, askyesno

book = openpyxl.open('Слова.xlsx',read_only=True)
sheet = book.active

## -- Выбор последовательности слов из таблички

n = list(range(2, sheet.max_row + 1))
#print(n)
random.shuffle(n)
#print(n)
ran = n[0]
selection = 'Тренажер'
# -- Объявление переменных

perevod1 = sheet[ran][1].value
perevod2 = sheet[ran][2].value
perevod3 = sheet[ran][3].value
count = 0
k = 0
s = 1

def next_word():
    global n, ran, perevod1, perevod2, perevod3, selection
    if len(n) > 1:
        n.pop(0)
        ran = n[0]
        res.set('')
        slovo.set(sheet[ran][0].value)
        perevod1 = sheet[ran][1].value
        perevod2 = sheet[ran][2].value
        perevod3 = sheet[ran][3].value
def restart_game():
    global n, ran, perevod1, perevod2, perevod3, count, k, selection, selT,selD
    if selection == 'Тренажер':
        n = list(range(2, sheet.max_row + 1))
        ran = n[0]
        VvodSlova.config(state='normal')
        VvodSlova.delete(0, END)
    elif selection == 'Стрик':
        selD = int(difCombobox.get())
        n = [i for i in range(2, sheet.max_row + 1) if sheet[i][5].value == selD]
        VvodSlova.config(state='normal')
        VvodSlova.delete(0, END)
    elif selection == 'Темы':
        selT = str(themCombobox.get()).strip().lower()
        n = [i for i in range(2, sheet.max_row + 1) if sheet[i][4].value == selT]
        VvodSlova.config(state='normal')
        VvodSlova.delete(0, END)
    random.shuffle(n)
    ran = n[0]
    res.set('')
    slovo.set(sheet[ran][0].value)
    perevod1 = sheet[ran][1].value
    perevod2 = sheet[ran][2].value
    perevod3 = sheet[ran][3].value
    count = 0
    k = 0
# -- Режимы
def selected(event):
    global n, ran, perevod1, perevod2, perevod3, count, k, selection
    selection = combobox.get()
    print('Выбран режим:',selection)
    VvodSlova.config(state='normal')
    VvodSlova.delete(0, END)
    if selection == 'Тренажер':
        difCombobox.pack_forget()
        themCombobox.pack_forget()
        newword.config(state='normal')
        btn.config(state='normal')
        restart_game()
    if selection == 'Стрик':
        k = 0
        newword.config(state='disabled')
        btn.config(state='disabled')
        VvodSlova.config(state='disabled')
        difCombobox.pack(side=tk.LEFT, anchor=tk.NW, padx=10, pady=10)
        themCombobox.pack_forget()
    if selection == 'Темы':
        k = 0
        newword.config(state='disabled')
        btn.config(state='disabled')
        VvodSlova.config(state='disabled')
        themCombobox.pack(side=tk.LEFT, anchor=tk.NW, padx=10, pady=10)
        difCombobox.pack_forget()
    if selection == 'Автор':
        k = 0
        VvodSlova.config(state='disabled')
        showinfo(title="А вот и автор!",message="https://github.com/TereDim",
        detail="ИИТ, ИБД-25-03, телефон +79389164047")
# -- Темы
def selectedT(event):
    global n, ran, perevod1, perevod2, perevod3, count, k, selT, maxvords
    newword.config(state='normal')
    btn.config(state='normal')
    VvodSlova.config(state='normal')
    VvodSlova.delete(0, END)
    selT = str(themCombobox.get()).strip().lower()
    n = [i for i in range(2, sheet.max_row + 1) if sheet[i][4].value == selT]
    random.shuffle(n)
    ran = n[0]
    slovo.set(sheet[ran][0].value)
    perevod1 = sheet[ran][1].value
    perevod2 = sheet[ran][2].value
    perevod3 = sheet[ran][3].value
    k = 0
    res.set('')
    maxvords = len(n)
    print(n)
    print('Тема', selT)
# -- Сложность
def selectedD(event):
    global n, ran, perevod1, perevod2, perevod3, count, k, selD, maxvords
    newword.config(state='normal')
    btn.config(state='normal')
    VvodSlova.config(state='normal')
    VvodSlova.delete(0, END)
    selD = int(difCombobox.get())
    n = [i for i in range(2, sheet.max_row + 1) if sheet[i][5].value == selD]
    random.shuffle(n)
    ran = n[0]
    slovo.set(sheet[ran][0].value)
    perevod1 = sheet[ran][1].value
    perevod2 = sheet[ran][2].value
    perevod3 = sheet[ran][3].value
    k = 0
    res.set('')
    slovo.set(sheet[ran][0].value)
    maxvords = len(n)
    print(n)
    print('Сложность', selD)
    
## -- Следущее слово

def click_next():
    global perevod1, perevod2, perevod3, ran, count, k, selection, n, maxvords
    # Режим тренажер
    VvodSlova.config(state='normal')
    VvodSlova.delete(0, END)
    if selection == 'Тренажер':
        if len(n) > 1:
            VvodSlova.delete(0, END)
            next_word()
            if count == 1:
                k += 1
                count = 0
        else:
            if count == 1:
                k += 1
                count = 0
            # Показ результата - тренажер
            showinfo(title="Поздравляю!",message="Вы прошли тренажер!",
                detail=f"Ваш результат: {k}/{sheet.max_row - 1}")
            VvodSlova.config(state='disabled')
# Предложение начать заново
            restart = askyesno("Перезапуск", "Хотите пройти заново?")
            if restart:
                restart_game()
    # Режим стрик
    elif selection == 'Стрик':
        if count == 1:
            k += 1
            count = 0
            if len(n) > 1:
                VvodSlova.delete(0, END)
                next_word()
            else:
                # Показ результата - стрик
                showinfo(title="Поздравляю!",message="Вы прошли режим стрик!",
                    detail=f"Ваш результат: {k}/{maxvords}")
                VvodSlova.config(state='disabled')
                # Предложение начать заново
                restart = askyesno("Перезапуск", "Хотите пройти заново?")
                if restart:
                    restart_game()
        else:
            showinfo(title="Ошибка",message="Вы должны ответить!",)
    elif selection == 'Темы':
        if len(n) > 1:
            VvodSlova.delete(0, END)
            next_word()
            if count == 1:
                k += 1
                count = 0
        else:
            if count == 1:
                k += 1
                count = 0
            # Показ результата - Темы
            showinfo(title="Поздравляю!",message="Вы прошли тему!",
                detail=f"Ваш результат: {k}/{maxvords}")
            VvodSlova.config(state='disabled')
            # Предложение начать заново
            restart = askyesno("Перезапуск", "Хотите пройти заново?")
            if restart:
                restart_game()
        
## -- ~~ Раздел - Окно

def click_button():
    global perevod1, perevod2, perevod3, count, k, selection
    #Режим тренажер и стрик
    if selection != 'Автор':
        otv = otvet.get().strip().lower().replace('ё','е')
        print('')
        print("Ответ пользователя:", otv)
        print("Правильные варианты:", perevod1, perevod2, perevod3)
        # Обработка слов (нижний регистр)
        p1 = str(perevod1).strip().lower()
        p2 = str(perevod2).strip().lower()
        p3 = str(perevod3).strip().lower()
        VvodSlova.config(state='disabled')
        #сравнение ответов
        if (otv == p1 or otv == p2 or otv == p3) and otv != 'none' and otv != 'None':
            print('Верно')
            res.set("Верно!")
            count = 1
        else:
            print('Неверно')
            res.set(perevod1)
            print(selection)
            if selection == 'Стрик':
                showinfo(title="Увы!", message="Вы ответили неверно!", 
                detail=f"Ваш результат:{k}/{maxvords}")
                restart_game()

def on_enter(event):
    click_button()

root = Tk()
root['bg'] = '#343c47'
root.title('Проект 2.0')
root.geometry('600x800')
root.resizable(width=False, height=False)

mode = ['Тренажер', 'Стрик', 'Темы', 'Автор']
mode_var = StringVar(value=mode[0])
combobox = ttk.Combobox(textvariable=mode_var, values=mode, state='readonly')
combobox.pack(side=tk.LEFT, anchor=tk.NW, padx=5, pady=10)
combobox.bind("<<ComboboxSelected>>", selected)

difficult = ['1', '2']
difCombobox = ttk.Combobox(values=difficult,state='readonly')
difCombobox.bind("<<ComboboxSelected>>", selectedD)

themes = ['Простые', 'Книги', 'Еда', 'Животные', 'Мебель', 'Одежда', \
'Природа', 'Развлечения', 'Профессии']
themCombobox = ttk.Combobox(values=themes, state='readonly')
themCombobox.bind("<<ComboboxSelected>>", selectedT)

slovo = StringVar()
slovo.set(sheet[ran][0].value)
otvet = StringVar()
res = StringVar()

label = ttk.Label(root,textvariable=slovo,font=("Arial", 14))
label.place(width=150,rely=0.4,relx=0.38, )

btn = Button(root,text="Ответить",command = click_button)
btn.place(x=480, y=590,height=28)

newword = Button(root,text="Далее",command = click_next)
newword.place(x=480, y=500,height=28)

VvodSlova = Entry(root, bg='white', width=35,font=("Arial", 14),textvariable=otvet)
VvodSlova.place(x=68, y=590)
root.bind('<Return>', on_enter)

check_label = ttk.Label(textvariable=res)
check_label.place(x=10, y=770)

root.mainloop()